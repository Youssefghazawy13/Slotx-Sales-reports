import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import zipfile
from io import BytesIO
from datetime import datetime

st.set_page_config(
    page_title="Slotx Sales & Inventory Reports",
    page_icon="📊",
    layout="centered"
)

def auto_fit_columns(ws):
    """Auto-fit all columns in the worksheet based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try: 
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def clean_brand_names(df):
    """Clean brand names:  remove extra spaces and normalize case"""
    if 'brand' in df.columns:
        df['brand'] = df['brand'].astype(str).str.strip().str.title()
    return df

def get_brand_deal_text(deal_percentage, rent_amount):
    """Generate brand deal text based on percentage and rent"""
    has_percentage = deal_percentage > 0
    has_rent = rent_amount > 0
    
    if has_rent and has_percentage:
        return f"{rent_amount} EGP + {deal_percentage}% Deducted From The Sales"
    elif has_rent: 
        return f"{rent_amount} EGP"
    elif has_percentage: 
        return f"{deal_percentage}% Deducted From The Sales"
    else:
        return ""

def remove_refunds_and_original_sales(sales_df):
    """Remove refund transactions AND their corresponding original sales"""
    if 'quantity' not in sales_df.columns or 'barcode' not in sales_df. columns:
        return sales_df, 0, 0
    
    original_count = len(sales_df)
    refunds = sales_df[sales_df['quantity'] < 0]. copy()
    
    if len(refunds) == 0:
        return sales_df, 0, 0
    
    refund_count = len(refunds)
    indices_to_remove = set()
    indices_to_remove.update(refunds.index. tolist())
    
    for idx, refund_row in refunds.iterrows():
        barcode = refund_row. get('barcode')
        refund_qty = abs(refund_row. get('quantity', 0))
        brand = refund_row.get('brand')
        
        potential_originals = sales_df[
            (sales_df['barcode'] == barcode) &
            (sales_df['brand'] == brand) &
            (sales_df['quantity'] > 0) &
            (~sales_df. index.isin(indices_to_remove))
        ]
        
        matching_sales = potential_originals[potential_originals['quantity'] == refund_qty]
        
        if len(matching_sales) > 0:
            indices_to_remove.add(matching_sales.index[0])
    
    cleaned_df = sales_df[~sales_df.index.isin(indices_to_remove)].copy()
    removed_count = original_count - len(cleaned_df)
    
    return cleaned_df, refund_count, removed_count

def get_best_selling_size(sales_data):
    """Extract and find the best selling size from product names"""
    if len(sales_data) == 0 or 'name_ar' not in sales_data.columns:
        return ''
    
    size_sales = {}
    
    for _, row in sales_data.iterrows():
        product_name = str(row. get('name_ar', ''))
        quantity = row.get('quantity', 0)
        
        if '-' in product_name:
            size = product_name.split('-')[-1]. strip()
            if size:
                size_sales[size] = size_sales.get(size, 0) + quantity
    
    if size_sales:
        best_size = max(size_sales, key=size_sales.get)
        return best_size
    
    return ''

def get_best_selling_products(sales_data):
    """Find the best selling product(s)"""
    if len(sales_data) == 0 or 'name_ar' not in sales_data.columns:
        return ''
    
    product_sales = {}
    
    for _, row in sales_data. iterrows():
        product_name = str(row.get('name_ar', ''))
        quantity = row.get('quantity', 0)
        
        if product_name:
            product_sales[product_name] = product_sales.get(product_name, 0) + quantity
    
    if not product_sales:
        return ''
    
    max_sales = max(product_sales.values())
    best_products = [product for product, sales in product_sales.items() if sales == max_sales]
    
    if len(best_products) == 1:
        return best_products[0]
    else:
        return ', '.join(best_products)

def create_sales_details_sheet(wb, brand_name, sales_data):
    """Create Sales Details sheet for a specific brand"""
    ws = wb.create_sheet(f"{brand_name} Sales Details")
    
    headers = ['Branch Name', 'Brand Name', 'Product Name', 'Barcode', 'Quantity', 'Price']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    total_quantity = 0
    total_price = 0
    
    for _, row in sales_data.iterrows():
        ws.append([
            row. get('branch_name', ''),
            row.get('brand', ''),
            row.get('name_ar', ''),
            row.get('barcode', ''),
            row.get('quantity', 0),
            row.get('total', 0)
        ])
        total_quantity += row.get('quantity', 0)
        total_price += row.get('total', 0)
    
    total_row = ['', '', '', '', f'Total={total_quantity}', f'Total={total_price}']
    ws.append(total_row)
    
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    
    auto_fit_columns(ws)

def create_inventory_sheet(wb, brand_name, inventory_data):
    """Create Inventory sheet for a specific brand"""
    ws = wb.create_sheet(f"{brand_name} Inventory")
    
    headers = ['Branch Name', 'Brand', 'Product Name', 'Barcodes', 'Product Price', 'Available Quantity']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for _, row in inventory_data.iterrows():
        ws.append([
            row.get('branch_name', ''),
            row.get('brand', ''),
            row.get('name_en', ''),
            row.get('barcodes', ''),
            row.get('sale_price', 0),
            row.get('available_quantity', 0)
        ])
    
    auto_fit_columns(ws)

def create_report_sheet(wb, brand_name, sales_data, inventory_data, payout_cycle, brand_settings):
    """Create Report sheet for a specific brand"""
    ws = wb.create_sheet(f"{brand_name} Report")
    
    branch_name = sales_data. iloc[0].get('branch_name', '') if len(sales_data) > 0 else ''
    
    # Calculate totals
    total_inventory_qty = inventory_data.get('available_quantity', pd.Series([0])).sum()
    total_inventory_value = (inventory_data.get('available_quantity', pd.Series([0])) * 
                            inventory_data.get('sale_price', pd.Series([0]))).sum()
    total_sales_qty = sales_data.get('quantity', pd.Series([0])).sum()
    total_sales_money = sales_data.get('total', pd.Series([0])).sum()
    
    # Get brand settings
    deal_percentage = brand_settings.get('deal_percentage', 0)
    rent_amount = brand_settings.get('rent_amount', 0)
    
    # Generate brand deal text
    brand_deal_text = get_brand_deal_text(deal_percentage, rent_amount)
    
    # Calculate after percentage and rent
    total_after_percentage = total_sales_money - (total_sales_money * deal_percentage / 100)
    total_after_rent = total_after_percentage - rent_amount
    
    # Get best selling info
    best_size = get_best_selling_size(sales_data)
    best_products = get_best_selling_products(sales_data)
    
    # Build report data
    report_data = [
        ['Branch Name:', branch_name],
        ['', ''],
        ['Brand Name:', brand_name],
        ['', ''],
        ['Brand Deal:', brand_deal_text],
        ['', ''],
        ['Payout Period:', payout_cycle],
        ['', ''],
        ['Best Selling Size:', best_size],
        ['Best Selling Product:', best_products],
        ['', ''],
        ['Total Brand Inventory Quantities:', total_inventory_qty],
        ['Total Brand Inventory Stock Price:', total_inventory_value],
        ['', ''],
        ['Total Sales (Products Quantities):', total_sales_qty],
        ['Total sales (Money):', total_sales_money],
        ['Total Sales After Percentage:', total_after_percentage],
        ['Total Sales After Rent:', total_after_rent]
    ]
    
    for row_data in report_data:
        ws.append(row_data)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                cell.font = Font(bold=True)
    
    auto_fit_columns(ws)

def create_all_brands_summary(sales_df, inventory_df, brand_settings_dict, payout_cycle):
    """Create a summary Excel file for all brands combined"""
    
    wb = Workbook()
    
    # Add metadata
    wb.properties.creator = "Slotx Reports Generator"
    wb.properties.lastModifiedBy = "Slotx Reports Generator"
    wb. properties.created = datetime.now()
    wb.properties.modified = datetime.now()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: All Sales Details
    ws_sales = wb. create_sheet("All Sales Details")
    headers_sales = ['Branch Name', 'Brand Name', 'Product Name', 'Barcode', 'Quantity', 'Price']
    ws_sales.append(headers_sales)
    
    for cell in ws_sales[1]:
        cell.font = Font(bold=True)
    
    total_sales_qty = 0
    total_sales_money = 0
    
    for _, row in sales_df.iterrows():
        ws_sales.append([
            row.get('branch_name', ''),
            row.get('brand', ''),
            row.get('name_ar', ''),
            row.get('barcode', ''),
            row.get('quantity', 0),
            row.get('total', 0)
        ])
        total_sales_qty += row.get('quantity', 0)
        total_sales_money += row.get('total', 0)
    
    # Add totals
    ws_sales. append(['', '', '', '', f'Total={total_sales_qty}', f'Total={total_sales_money}'])
    for cell in ws_sales[ws_sales.max_row]: 
        cell.font = Font(bold=True)
    
    auto_fit_columns(ws_sales)
    
    # Sheet 2: All Inventory
    ws_inventory = wb.create_sheet("All Inventory")
    headers_inventory = ['Branch Name', 'Brand', 'Product Name', 'Barcodes', 'Product Price', 'Available Quantity']
    ws_inventory.append(headers_inventory)
    
    for cell in ws_inventory[1]:
        cell.font = Font(bold=True)
    
    total_inventory_qty = 0
    total_inventory_value = 0
    
    for _, row in inventory_df.iterrows():
        qty = row.get('available_quantity', 0)
        price = row.get('sale_price', 0)
        ws_inventory.append([
            row.get('branch_name', ''),
            row.get('brand', ''),
            row.get('name_en', ''),
            row.get('barcodes', ''),
            price,
            qty
        ])
        total_inventory_qty += qty
        total_inventory_value += qty * price
    
    auto_fit_columns(ws_inventory)
    
    # Sheet 3: Brands Deals
    ws_deals = wb. create_sheet("Brands Deals")
    headers_deals = ['Brand Name', 'Deal Percentage (%)', 'Rent Amount (EGP)', 'Brand Deal']
    ws_deals.append(headers_deals)
    
    for cell in ws_deals[1]:
        cell.font = Font(bold=True)
    
    for brand in sorted(brand_settings_dict.keys()):
        settings = brand_settings_dict[brand]
        deal_text = get_brand_deal_text(settings['deal_percentage'], settings['rent_amount'])
        ws_deals.append([
            brand,
            settings['deal_percentage'],
            settings['rent_amount'],
            deal_text
        ])
    
    auto_fit_columns(ws_deals)
    
    # Sheet 4: Summary Report
    ws_report = wb. create_sheet("Summary Report")
    
    # Calculate best selling sizes (top 3)
    size_sales = {}
    for _, row in sales_df.iterrows():
        product_name = str(row.get('name_ar', ''))
        quantity = row.get('quantity', 0)
        if '-' in product_name:
            size = product_name.split('-')[-1].strip()
            if size:
                size_sales[size] = size_sales. get(size, 0) + quantity
    
    top_sizes = sorted(size_sales.items(), key=lambda x: x[1], reverse=True)[:3]
    best_sizes_text = ', '.join([f"{size} ({qty} units)" for size, qty in top_sizes]) if top_sizes else ''
    
    # Calculate best selling products (top 3)
    product_sales = {}
    for _, row in sales_df.iterrows():
        product_name = str(row.get('name_ar', ''))
        quantity = row. get('quantity', 0)
        if product_name:
            product_sales[product_name] = product_sales.get(product_name, 0) + quantity
    
    top_products = sorted(product_sales.items(), key=lambda x: x[1], reverse=True)[:3]
    best_products_text = ', '.join([f"{prod} ({qty} units)" for prod, qty in top_products]) if top_products else ''
    
    # Calculate totals with deductions
    total_percentage_deducted = 0
    total_rent_deducted = 0
    
    brands = sales_df['brand'].dropna().unique()
    for brand in brands:
        brand_sales = sales_df[sales_df['brand'] == brand]
        brand_total = brand_sales['total'].sum()
        
        settings = brand_settings_dict. get(brand, {'deal_percentage': 0, 'rent_amount': 0})
        
        percentage_deduction = brand_total * settings['deal_percentage'] / 100
        rent_deduction = settings['rent_amount']
        
        total_percentage_deducted += percentage_deduction
        total_rent_deducted += rent_deduction
    
    total_after_all_deductions = total_sales_money - total_percentage_deducted - total_rent_deducted
    
    # Build report
    report_data = [
        ['Payout Period:', payout_cycle],
        ['', ''],
        ['Total Sales (Money):', total_sales_money],
        ['Total Sales (Quantities):', total_sales_qty],
        ['', ''],
        ['Total Inventory Quantities:', total_inventory_qty],
        ['Total Inventory Value:', total_inventory_value],
        ['', ''],
        ['Best Selling Sizes (Top 3):', best_sizes_text],
        ['Best Selling Products (Top 3):', best_products_text],
        ['', ''],
        ['Total Percentage Deducted (Money):', total_percentage_deducted],
        ['Total Rent Deducted (Money):', total_rent_deducted],
        ['Total Sales After All Deductions:', total_after_all_deductions]
    ]
    
    for row_data in report_data: 
        ws_report.append(row_data)
    
    for row in ws_report.iter_rows(min_row=1, max_row=ws_report. max_row, min_col=1, max_col=1):
        for cell in row: 
            if cell.value:
                cell.font = Font(bold=True)
    
    auto_fit_columns(ws_report)
    
    return wb

def process_files(sales_df, inventory_df, payout_cycle, brand_settings_dict):
    """Process the sales and inventory files and generate brand reports"""
    
    sales_df. columns = sales_df.columns.str.strip()
    inventory_df.columns = inventory_df.columns.str.strip()
    
    sales_df = clean_brand_names(sales_df)
    inventory_df = clean_brand_names(inventory_df)
    
    sales_df = sales_df.dropna(how='all')
    inventory_df = inventory_df.dropna(how='all')
    
    # Remove refunds and get stats
    sales_df, refund_count, total_removed = remove_refunds_and_original_sales(sales_df)
    
    brands = sales_df['brand'].dropna().unique()
    
    # Show processing summary
    st.info(f"📊 **Processing Summary:**\n- **{len(brands)} brands** detected\n- **{refund_count} refunds** + **{total_removed - refund_count} original sales** removed ({total_removed} total transactions deleted)")
    
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Create individual brand files
        for brand in brands: 
            brand_sales = sales_df[sales_df['brand'] == brand]. copy()
            brand_inventory = inventory_df[inventory_df['brand'] == brand].copy()
            
            if len(brand_sales) == 0:
                continue
            
            # Get brand settings
            brand_settings = brand_settings_dict.get(brand, {'deal_percentage': 0, 'rent_amount': 0})
            
            # Create workbook with proper metadata
            wb = Workbook()
            
            # Add metadata to prevent Excel locked/protected view issues
            wb.properties.creator = "Slotx Reports Generator"
            wb.properties.lastModifiedBy = "Slotx Reports Generator"
            wb.properties.created = datetime.now()
            wb.properties.modified = datetime.now()
            
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            create_sales_details_sheet(wb, brand, brand_sales)
            create_inventory_sheet(wb, brand, brand_inventory)
            create_report_sheet(wb, brand, brand_sales, brand_inventory, payout_cycle, brand_settings)
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_data = excel_buffer.getvalue()
            
            safe_brand_name = str(brand).replace('/', '-').replace('\\', '-').replace(':', '-').strip()
            zip_file.writestr(f"{safe_brand_name}.xlsx", excel_data)
            
            excel_buffer.close()
            wb.close()
        
        # Create All Brands Summary file
        summary_wb = create_all_brands_summary(sales_df, inventory_df, brand_settings_dict, payout_cycle)
        summary_buffer = BytesIO()
        summary_wb.save(summary_buffer)
        summary_data = summary_buffer.getvalue()
        
        # Add summary to ZIP
        zip_file.writestr("All_Brands_Summary.xlsx", summary_data)
        
        summary_buffer.close()
        summary_wb.close()
    
    zip_buffer.seek(0)
    return zip_buffer

# Streamlit UI
st.title("📊 Slotx Sales & Inventory Reports Generator")
st.markdown("Generate brand-specific sales and inventory reports from Excel files")

st.divider()

# Payout Cycle
st.subheader("📅 Payout Cycle")
payout_cycle = st.selectbox(
    "Select Payout Cycle",
    options=["-- Select Payout Cycle --", "Payout Cycle 1", "Payout Cycle 2"],
    index=0,
    help="Choose the payout cycle for this report"
)

# Validate payout cycle selection
payout_cycle_selected = payout_cycle != "-- Select Payout Cycle --"

if not payout_cycle_selected: 
    st.warning("⚠️ Please select a Payout Cycle to continue")

st.divider()

# Sales file upload
st.subheader("📈 Sales Sheet")
sales_file = st.file_uploader(
    "Upload Sales Excel File",
    type=['xlsx', 'xls'],
    key='sales',
    help="Upload sales data to detect brands"
)

# Brand settings (show after sales upload)
brand_settings_dict = {}

if sales_file:
    try:
        # Read and process sales file
        sales_df_temp = pd.read_excel(sales_file)
        sales_df_temp. columns = sales_df_temp. columns.str.strip()
        sales_df_temp = clean_brand_names(sales_df_temp)
        sales_df_temp = sales_df_temp.dropna(how='all')
        sales_df_temp, _, _ = remove_refunds_and_original_sales(sales_df_temp)
        
        brands = sales_df_temp['brand'].dropna().unique()
        
        st.success(f"✅ Found {len(brands)} brand(s) in Sales data")
        
        st.divider()
        st.subheader("📊 Brand Settings")
        st.markdown("Enter deal percentage and/or rent amount for each brand:")
        st.info("💡 **Tip:** Leave at 0 if not applicable.  You can set percentage only, rent only, or both.")
        
        # Create form for each brand
        for brand in sorted(brands):
            with st.expander(f"🏷️ **{brand}**", expanded=True):
                col1, col2 = st. columns(2)
                
                with col1:
                    deal_percentage = st.number_input(
                        "Deal Percentage (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=0.0,
                        step=0.5,
                        key=f"deal_{brand}",
                        help="Percentage to deduct from total sales (leave 0 if no percentage)"
                    )
                
                with col2:
                    rent_amount = st.number_input(
                        "Rent Amount (EGP)",
                        min_value=0.0,
                        value=0.0,
                        step=100.0,
                        key=f"rent_{brand}",
                        help="Fixed rent amount to deduct (leave 0 if no rent)"
                    )
                
                # Show preview of brand deal text
                preview_text = get_brand_deal_text(deal_percentage, rent_amount)
                if preview_text:
                    st.caption(f"📝 Brand Deal will show: **{preview_text}**")
                else:
                    st.caption("📝 No deal configured for this brand")
                
                brand_settings_dict[brand] = {
                    'deal_percentage': deal_percentage,
                    'rent_amount': rent_amount
                }
        
    except Exception as e:
        st.error(f"❌ Error reading sales file: {str(e)}")

st.divider()

# Inventory file upload
st.subheader("📦 Inventory Sheet")
inventory_file = st.file_uploader(
    "Upload Inventory Excel File",
    type=['xlsx', 'xls'],
    key='inventory',
    help="Upload inventory data"
)

st.divider()

# Generate button
if payout_cycle_selected and sales_file and inventory_file and len(brand_settings_dict) > 0:
    if st.button("🚀 Generate Reports", type="primary", use_container_width=True):
        try:
            with st.spinner("Processing files...  Please wait"):
                sales_df = pd.read_excel(sales_file)
                inventory_df = pd.read_excel(inventory_file)
                
                zip_buffer = process_files(sales_df, inventory_df, payout_cycle, brand_settings_dict)
                
                brands_count = len(brand_settings_dict)
                
                st.success(f"✅ Successfully generated reports for {brands_count} brand(s) + All Brands Summary!")
                
                st.download_button(
                    label="📥 Download Brands Reports (ZIP)",
                    data=zip_buffer,
                    file_name=f"Brands_Reports_{payout_cycle. replace(' ', '_')}.zip",
                    mime="application/zip",
                    use_container_width=True
                )
                
        except Exception as e:
            st.error(f"❌ Error processing files: {str(e)}")
            st.exception(e)
else:
    if not payout_cycle_selected:
        st.info("ℹ️ Please select a Payout Cycle first")
    elif not sales_file:
        st.info("ℹ️ Please upload Sales Excel file")
    elif not inventory_file:
        st.info("ℹ️ Please upload Inventory Excel file")
    else:
        st.info("ℹ️ Please fill in brand settings above")
